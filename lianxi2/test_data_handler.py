# -*- coding: utf-8 -*-
# @Time    : 2021/3/1 9:34 下午
# @Author  : Lewin
# @FileName: test_data_handler.py
# @Software: PyCharm
from openpyxl import load_workbook


def get_data_from_excel(file, sheet_name=None):
    """
    从Excel中获取测试数据
    :param file:
    :param sheet_name:
    :return:
    """
    # 1.加载Excel文件
    wb = load_workbook(file)
    # 2.获取对应的表
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    # 3.获取几行几列
    row = ws.max_row
    col = ws.max_column
    # 4. 获取cell中的数据并组织成需要的结构
    # 创建一个空列表用来接收数据
    data = []
    # 先获取标题
    title = {}
    for i in range(1, col + 1):
        title[i] = ws.cell(row=1, column=i).value

    # title = {
    #     1: 'title',
    #     2: 'username',
    #     3: 'password',
    #     4: 'expect'
    # }
    # 获取值
    for j in range(2, row + 1):  # 获取行
        temp = {}
        for i in range(1, col + 1):  # 获取列
            temp[title[i]] = ws.cell(row=j, column=i).value

        data.append(temp)
    return data


if __name__ == '__main__':
    res = get_data_from_excel('testcases.xlsx', sheet_name='register')
    print(res)
