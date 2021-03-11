# -*- coding: utf-8 -*-
# @Time    : 2021/2/22 3:54 下午
# @Author  : Lewin
# @FileName: excel.py
# @Software: PyCharm
# from openpyxl import load_workbook
# wb = load_workbook('cases.xlsx')
#
# for sheet in wb:
#
#     print(sheet)
#
#
# ws = wb['login']
# # print(ws)
# #
# # print('---------')
# #
# # print(wb.active)
# #
# #
# # print(ws, type(ws))
#
# print('总共有：{}行，{}列'.format(ws.max_row, ws.max_column))
#
# cell = ws.cell(row=2, column=1)
# print(cell, cell.value, type(cell))

from openpyxl import load_workbook


def get_data_from_excel(file, sheet_name=None):
    """
    从excel文件中获取测试用例数据
    """
    # 1. 加载excel文件
    wb = load_workbook(file)
    # 2. 获取对应的表
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    # 获取几行几列
    row = ws.max_row
    column = ws.max_column
    # 3. 获取cell中的数据并组织成需要的结构
    # 创建一个空列表用来接收数据
    data = []
    # 先获取标题
    title = {}
    for i in range(1, column + 1):
        title[i] = ws.cell(row=1, column=i).value
        print(title[i])
    # 获取值
    for j in range(2, row+1):
        temp = {}
        for i in range(1, column + 1):
            temp[title[i]] = ws.cell(row=j, column=i).value
            print(temp[title[i]])
        temp[title[i]] = eval(temp[title[i]])

        data.append(temp)
    return data


if __name__ == '__main__':
    res = get_data_from_excel('cases.xlsx', sheet_name='login')
    print(res)
    print(res[0]['expect'], type(res[0]['expect']))



