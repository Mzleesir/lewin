# -*- coding: utf-8 -*-
# @Time    : 2021/3/9 9:11 下午
# @Author  : Lewin
# @FileName: 操作excel.py
# @Software: PyCharm

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook, Worksheet
from openpyxl.cell.cell import Cell

wb = load_workbook('cases.xlsx')
print(wb)
print('------------')

for sh in wb:
    print(sh)

print('------------')

ws = wb['login']
print(ws)

print('------ws------')

cell = ws.cell(row=1, column=2)
print(cell, cell.value)

print('------------')

print('总共有{}行, {}列'.format(ws.max_row,))


