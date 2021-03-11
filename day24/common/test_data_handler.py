#! /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020/11/13 21:49
# @Author : 心蓝
from openpyxl import load_workbook


def get_data_from_excel(file, sheet_name=None):
    """
    从excel文件中获取测试用例数据
    :param file: excel文件名
    :param sheet_name: 表名
    :return:
    """
    # 1. 加载excel文件
    wb = load_workbook(file)
    # 2. 获取对应的表
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    # 获取几行几列
    row = ws.max_row
    column = ws.max_column
    # 3. 获取cell中的数据并组织成需要的结构
    # 创建一个空列表用来接收数据
    data = []
    # 先获取标题
    title = {}
    for i in range(1, column+1):
        title[i] = ws.cell(row=1, column=i).value
    # title = {i: ws.cell(row=1, column=i).value for i in range(1, column+1)}
        # print(ws.cell(row=1, column=i).value)
        # print(1, i)
    # print(title)
    # title = {
    #     1: 'title',
    #     2: 'username',
    #     3: 'password',
    #     4: 'expect'
    # }
    # 获取值
    # print(row, column)
    for j in range(2, row+1):
        # 临时变量，表示一行数据
        temp = {}
        for i in range(1, column+1):
            # print(j, i)

            # print(title[i], ws.cell(row=j, column=i).value, end=' ')
            temp[title[i]] = ws.cell(row=j, column=i).value
        # 循环结束的时候i=column 也就是最后一列 title[i]取出expect
        # print(temp)
        # print(temp[title[i]])
        # temp[title[i]] = eval(temp[title[i]])
        # print(temp)
        data.append(temp)

    return data


if __name__ == '__main__':
    res = get_data_from_excel('../testdata/testcases.xlsx', sheet_name='register')
    print(res)
    # print(res[0]['expect'], type(res[0]['expect']))