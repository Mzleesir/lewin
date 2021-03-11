#! /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020/11/13 21:28
# @Author : 心蓝
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook, Worksheet
from openpyxl.cell.cell import Cell
# 加载工作簿，一个excel文件
# 封装为一个openpyxl.workbook.workbook.Workbook的对象
wb = load_workbook('cases.xlsx')
print(wb, type(wb))
for sheet in wb:
    print(sheet)
# 读取表
# 表又被封装为Worksheet对象
# 通过表名获取表对象，类似字典取值的方式
# wb[表名]
ws = wb['login']
# wb.active 获取当前表
print('---------------')
print(wb.active)

print(ws, type(ws))
print('总共有：{}行，{}列'.format(ws.max_row, ws.max_column))
# 可以获取总共有几行几列

# 获取单元格
# 根据行row列column取单元格，行列都是从1开始
cell = ws.cell(row=7, column=1)
# 通过cell对象的value属性可以获取值
# value只有两周那个类型，字符串，数值
# 如果cell里没有值，返回None
print(cell, cell.value, type(cell))
# Worksheet().max_column
#

